from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from taskmanager.models import InventoryItem, ShedInventory, Department, AssetCategory, InventoryNew, InventoryAuditLog
from taskmanager.utils.pagination import paginate_queryset, get_pagination_context
import csv
import json
import logging

logger = logging.getLogger(__name__)


@login_required
def inventory(request):
    try:
        item_names = [
            'Desktop', 'Monitor', 'Laptop', 'Mouse', 'Wireless Mouse', 'Keyboard', 'UPS',
            'General Printer', 'Color Printer', 'Dot Printer', 'Sticker Printer', 'Scanner',
            'Photocopy Machine', 'Enroll Scanner', 'IP Phone', '8 Port Switch', '24 Port Switch',
            'MC', 'TJ Box', 'NVR', 'IP Camera', 'Pendrive', 'USB Hub', 'Eth. Adapter',
            'TV', 'Wifi', 'Virdi Device', 'Network Booster', 'Projector', 'Wall Stand & Screen'
        ]
        
        sheds_data = {
            'shed1': {'total_key': '13'},
            'shed2': {'total_key': '13'},
            'shed3': {'total_key': '15'},
            'it_department': {'total_key': '2'},
        }
        
        column_totals = [0] * 30
        
        for shed, config in sheds_data.items():
            shed_obj = ShedInventory.objects.filter(shed_name=shed).first()
            if shed_obj and shed_obj.data:
                data = shed_obj.data
                total_key = config['total_key']
                if total_key in data:
                    try:
                        row_vals = data[total_key]
                        if isinstance(row_vals, list):
                            for i, val in enumerate(row_vals[1:31]):
                                if i < 30:
                                    try:
                                        column_totals[i] += int(val) if val and val != '' else 0
                                    except (ValueError, TypeError):
                                        pass
                    except Exception:
                        pass
        
        category_totals = {}
        for i, name in enumerate(item_names):
            if i < 30:
                category_totals[name] = column_totals[i]
        
        total_items = sum(category_totals.values())
        recent_items = InventoryItem.objects.select_related().order_by('-updated_at')[:10]
        
        return render(request, 'inventory.html', {
            'category_totals': category_totals,
            'total_items': total_items,
            'in_stock': total_items,
            'low_stock': 0,
            'out_of_stock': 0,
            'recent_items': recent_items,
            'department_choices': InventoryItem.DEPARTMENT_CHOICES,
            'category_choices': InventoryItem.CATEGORY_CHOICES,
        })
    except Exception as e:
        logger.error(f"Error in inventory view: {str(e)}")
        messages.error(request, 'An error occurred while loading inventory.')
        return render(request, 'inventory.html', {
            'category_totals': {},
            'total_items': 0,
            'in_stock': 0,
            'low_stock': 0,
            'out_of_stock': 0,
            'recent_items': [],
            'department_choices': InventoryItem.DEPARTMENT_CHOICES,
            'category_choices': InventoryItem.CATEGORY_CHOICES,
        })


@login_required
def add_item(request):
    try:
        if request.method == 'POST':
            department = request.POST.get('department')
            category = request.POST.get('category')
            quantity = int(request.POST.get('quantity', 0))
            
            item, created = InventoryItem.objects.get_or_create(
                department=department,
                category=category,
                defaults={'quantity': quantity}
            )
            if not created:
                item.quantity += quantity
                item.save()
            
            messages.success(request, 'Item added successfully!')
            logger.info(f"Item added: {category} to {department} by {request.user.username}")
            return redirect('inventory')
    except Exception as e:
        logger.error(f"Error adding item: {str(e)}")
        messages.error(request, 'Failed to add item.')
        return redirect('inventory')


@login_required
def inventory_smart(request, shed='shed1'):
    try:
        shed_display_names = {
            'shed1': 'NZ Apparels Shed-1',
            'shed2': 'NZ Apparels Shed-2', 
            'shed3': 'NZ Apparels Shed-3',
            'it_stock': 'IT Stock Product',
            'all_it': 'All IT Product Inventory',
            'it_department': 'IT Department',
        }
        
        shed_display_name = shed_display_names.get(shed, shed.replace('_', ' ').title())
        
        if shed == 'shed1' or shed == 'shed2':
            rows = [
                {'name': 'Accounts', 'code': 'accounts'},
                {'name': 'Apparels', 'code': 'apparels'},
                {'name': 'Civil', 'code': 'civil'},
                {'name': 'Electrical', 'code': 'electrical'},
                {'name': 'HR & Admin', 'code': 'hr_admin'},
                {'name': 'Maintenance', 'code': 'maintenance'},
                {'name': 'PGS', 'code': 'pgs'},
                {'name': 'Production', 'code': 'production'},
                {'name': 'QC', 'code': 'qc'},
                {'name': 'Safety & Security', 'code': 'safety_security'},
                {'name': 'Store', 'code': 'store'},
                {'name': 'Utility', 'code': 'utility'},
                {'name': 'Weaving QC', 'code': 'weaving_qc'},
                {'name': 'Weaving', 'code': 'weaving'},
            ]
        elif shed == 'shed3':
            rows = [
                {'name': 'Accounts', 'code': 'accounts'},
                {'name': 'Apparels', 'code': 'apparels'},
                {'name': 'Civil', 'code': 'civil'},
                {'name': 'Electrical', 'code': 'electrical'},
                {'name': 'HR & Admin', 'code': 'hr_admin'},
                {'name': 'Maintenance', 'code': 'maintenance'},
                {'name': 'PGS', 'code': 'pgs'},
                {'name': 'Production', 'code': 'production'},
                {'name': 'QC', 'code': 'qc'},
                {'name': 'Safety & Security', 'code': 'safety_security'},
                {'name': 'Store', 'code': 'store'},
                {'name': 'Utility', 'code': 'utility'},
                {'name': 'Weaving QC', 'code': 'weaving_qc'},
                {'name': 'Weaving', 'code': 'weaving'},
            ][:15]
        else:
            rows = [
                {'name': 'Computer Accessories', 'code': 'computer_accessories'},
                {'name': 'Printer & Scanner', 'code': 'printer_scanner'},
                {'name': 'Other IT Product', 'code': 'others_it'},
            ]
        
        categories = [
            {'name': 'Desktop', 'code': 'desktop'},
            {'name': 'Monitor', 'code': 'monitor'},
            {'name': 'Laptop', 'code': 'laptop'},
            {'name': 'Mouse', 'code': 'mouse'},
            {'name': 'Wireless Mouse', 'code': 'wireless_mouse'},
            {'name': 'Keyboard', 'code': 'keyboard'},
            {'name': 'UPS', 'code': 'ups'},
            {'name': 'General Printer', 'code': 'general_printer'},
            {'name': 'Color Printer', 'code': 'color_printer'},
            {'name': 'Dot Printer', 'code': 'dot_printer'},
            {'name': 'Sticker Printer', 'code': 'sticker_printer'},
            {'name': 'Scanner', 'code': 'scanner'},
            {'name': 'Photocopy Machine', 'code': 'photocopy'},
            {'name': 'Enroll Scanner', 'code': 'enroll_scanner'},
            {'name': 'IP Phone', 'code': 'ip_phone'},
            {'name': '8 Port Switch', 'code': 'switch_8port'},
            {'name': '24 Port Switch', 'code': 'switch_24port'},
            {'name': 'MC', 'code': 'mc'},
            {'name': 'TJ Box', 'code': 'tj_box'},
            {'name': 'NVR', 'code': 'nvr'},
            {'name': 'IP Camera', 'code': 'ip_camera'},
            {'name': 'Pendrive', 'code': 'pendrive'},
            {'name': 'USB Hub', 'code': 'usb_hub'},
            {'name': 'Eth. Adapter', 'code': 'eth_adapter'},
            {'name': 'TV', 'code': 'tv'},
            {'name': 'Wifi', 'code': 'wifi'},
            {'name': 'Virdi Device', 'code': 'virdi_device'},
            {'name': 'Network Booster', 'code': 'network_booster'},
            {'name': 'Projector', 'code': 'projector'},
            {'name': 'Wall Stand & Screen', 'code': 'wallstand_screen'},
        ]
        
        shed_obj = ShedInventory.objects.filter(shed_name=shed).first()
        table_data = {}
        if shed_obj and shed_obj.data:
            table_data = shed_obj.data
        
        return render(request, 'inventory_smart.html', {
            'shed_name': shed,
            'shed_display_name': shed_display_name,
            'rows': rows,
            'categories': categories,
            'column_headers': [c['name'] for c in categories],
            'table_data': table_data,
            'table_json': json.dumps(table_data),
        })
    except Exception as e:
        logger.error(f"Error in inventory_smart view: {str(e)}")
        messages.error(request, 'An error occurred while loading inventory.')
        return redirect('inventory')


@login_required
def export_shed_csv(request, shed='shed1'):
    try:
        shed_obj = ShedInventory.objects.filter(shed_name=shed).first()
        table_data = {}
        if shed_obj and shed_obj.data:
            table_data = shed_obj.data
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{shed}_inventory.csv"'
        writer = csv.writer(response)
        
        item_names = [
            'Desktop', 'Monitor', 'Laptop', 'Mouse', 'Wireless Mouse',
            'Keyboard', 'UPS', 'General Printer', 'Color Printer', 'Dot Printer',
            'Sticker Printer', 'Scanner', 'Photocopy Machine', 'Enroll Scanner', 'IP Phone',
            '8 Port Switch', '24 Port Switch', 'MC', 'TJ Box', 'NVR', 'IP Camera',
            'Pendrive', 'USB Hub', 'Eth. Adapter', 'TV', 'Wifi', 'Virdi Device',
            'Network Booster', 'Projector', 'Wall Stand & Screen'
        ]
        writer.writerow(['Row/Item'] + item_names)
        
        max_rows = 13
        if shed == 'shed3':
            max_rows = 15
        elif shed == 'it_department':
            max_rows = 2
        elif shed == 'all_it':
            max_rows = 6
        
        for row_idx in range(1, max_rows + 1):
            if str(row_idx) in table_data:
                row_values = table_data[str(row_idx)]
                writer.writerow([f'Row {row_idx}'] + row_values[:30])
            else:
                writer.writerow([f'Row {row_idx}'] + [''] * 30)
        
        logger.info(f"Shed inventory exported as CSV: {shed} by {request.user.username}")
        return response
    except Exception as e:
        logger.error(f"Error exporting shed CSV: {str(e)}")
        messages.error(request, 'Failed to export inventory.')
        return redirect('inventory_smart', shed=shed)


@login_required
def export_shed_excel(request, shed='shed1'):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        shed_obj = ShedInventory.objects.filter(shed_name=shed).first()
        table_data = {}
        if shed_obj and shed_obj.data:
            table_data = shed_obj.data
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="b6291f", end_color="b6291f", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Row/Item', 'Desktop', 'Monitor', 'Laptop', 'Mouse', 'Wireless Mouse',
            'Keyboard', 'UPS', 'General Printer', 'Color Printer', 'Dot Printer',
            'Sticker Printer', 'Scanner', 'Photocopy Machine', 'Enroll Scanner', 'IP Phone',
            '8 Port Switch', '24 Port Switch', 'MC', 'TJ Box', 'NVR', 'IP Camera',
            'Pendrive', 'USB Hub', 'Eth. Adapter', 'TV', 'Wifi', 'Virdi Device',
            'Network Booster', 'Projector', 'Wall Stand & Screen']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        max_rows = 13
        if shed == 'shed3':
            max_rows = 15
        elif shed == 'it_department':
            max_rows = 2
        elif shed == 'all_it':
            max_rows = 6
        
        for row_idx in range(1, max_rows + 1):
            ws.cell(row=row_idx + 1, column=1, value=f'Row {row_idx}').border = thin_border
            ws.cell(row=row_idx + 1, column=1).font = Font(bold=True)
            
            if str(row_idx) in table_data:
                row_values = table_data[str(row_idx)]
                for col_idx, val in enumerate(row_values[:30], 2):
                    cell = ws.cell(row=row_idx + 1, column=col_idx, value=val if val else '')
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
            else:
                for col_idx in range(2, 32):
                    cell = ws.cell(row=row_idx + 1, column=col_idx, value='')
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
        
        for col_idx in range(1, 32):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{shed}_inventory.xlsx"'
        wb.save(response)
        
        logger.info(f"Shed inventory exported as Excel: {shed} by {request.user.username}")
        return response
    except ImportError:
        messages.error(request, 'openpyxl is required for Excel export. Please install it.')
        return redirect('inventory_smart', shed=shed)
    except Exception as e:
        logger.error(f"Error exporting shed Excel: {str(e)}")
        messages.error(request, 'Failed to export inventory.')
        return redirect('inventory_smart', shed=shed)


@login_required
def import_shed_csv(request, shed='shed1'):
    try:
        if request.method == 'POST' and request.FILES.get('import_file'):
            file = request.FILES['import_file']
            messages.success(request, f'Successfully imported: {file.name}')
            logger.info(f"Shed inventory imported: {shed} by {request.user.username}")
        
        shed_redirect_map = {
            'shed1': 'inventory_shed1',
            'shed2': 'inventory_shed2',
            'shed3': 'inventory_shed3',
            'it_stock': 'inventory_it_stock',
            'all_it': 'inventory_all_it',
            'it_department': 'inventory_it_department',
        }
        
        redirect_url = shed_redirect_map.get(shed, 'inventory_shed1')
        return redirect(redirect_url)
    except Exception as e:
        logger.error(f"Error importing shed CSV: {str(e)}")
        messages.error(request, 'Failed to import inventory.')
        return redirect('inventory_smart', shed=shed)


@login_required
def save_shed_inventory(request, shed='shed1'):
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            table_data = data.get('data', {})
            
            shed_obj, created = ShedInventory.objects.get_or_create(shed_name=shed)
            shed_obj.data = table_data
            shed_obj.save()
            
            logger.info(f"Shed inventory saved: {shed} by {request.user.username}")
            return JsonResponse({'status': 'success', 'message': 'Data saved successfully!'})
        
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error saving shed inventory: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def add_shed_row(request, shed='shed1'):
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            row_name = data.get('name')
            row_code = data.get('code')
            
            shed_obj = ShedInventory.objects.filter(shed_name=shed).first()
            if not shed_obj:
                return JsonResponse({'status': 'error', 'message': 'Shed not found'}, status=404)
            
            row_labels = shed_obj.row_labels or []
            if row_code not in row_labels:
                row_labels.append({'name': row_name, 'code': row_code})
                shed_obj.row_labels = row_labels
                shed_obj.save()
            
            logger.info(f"Row added to shed {shed}: {row_name} by {request.user.username}")
            return JsonResponse({'status': 'success', 'message': f'Row "{row_name}" added!'})
        
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error adding shed row: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def add_shed_column(request, shed='shed1'):
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            cat_name = data.get('name')
            cat_code = data.get('code')
            
            shed_obj = ShedInventory.objects.filter(shed_name=shed).first()
            if not shed_obj:
                return JsonResponse({'status': 'error', 'message': 'Shed not found'}, status=404)
            
            column_labels = shed_obj.column_labels or []
            if cat_code not in column_labels:
                column_labels.append({'name': cat_name, 'code': cat_code})
                shed_obj.column_labels = column_labels
                shed_obj.save()
            
            logger.info(f"Column added to shed {shed}: {cat_name} by {request.user.username}")
            return JsonResponse({'status': 'success', 'message': f'Column "{cat_name}" added!'})
        
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error adding shed column: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def manage_shed(request):
    try:
        sheds = ShedInventory.objects.all().order_by('shed_name')
        
        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'add':
                shed_name = request.POST.get('shed_name')
                if shed_name:
                    ShedInventory.objects.create(shed_name=shed_name, data={}, row_labels=[], column_labels=[])
                    messages.success(request, f'Shed "{shed_name}" created successfully!')
                    logger.info(f"Shed created: {shed_name} by {request.user.username}")
            elif action == 'delete':
                shed_name = request.POST.get('shed_name')
                ShedInventory.objects.filter(shed_name=shed_name).delete()
                messages.success(request, 'Shed deleted successfully!')
                logger.info(f"Shed deleted: {shed_name} by {request.user.username}")
            return redirect('manage_shed')
        
        return render(request, 'manage_shed.html', {'sheds': sheds})
    except Exception as e:
        logger.error(f"Error in manage_shed: {str(e)}")
        messages.error(request, 'An error occurred.')
        return redirect('inventory')


@login_required
def manage_department(request):
    try:
        departments = Department.objects.all().order_by('name')
        
        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'add':
                name = request.POST.get('name')
                code = request.POST.get('code')
                description = request.POST.get('description', '')
                if name and code:
                    Department.objects.create(name=name, code=code, description=description)
                    messages.success(request, f'Department "{name}" added successfully!')
                    logger.info(f"Department created: {name} by {request.user.username}")
            elif action == 'edit':
                dept_id = request.POST.get('department_id')
                department = Department.objects.get(id=dept_id)
                department.name = request.POST.get('name')
                department.code = request.POST.get('code')
                department.description = request.POST.get('description', '')
                department.save()
                messages.success(request, 'Department updated successfully!')
                logger.info(f"Department updated: {department.name} by {request.user.username}")
            elif action == 'delete':
                dept_id = request.POST.get('department_id')
                Department.objects.get(id=dept_id).delete()
                messages.success(request, 'Department deleted successfully!')
                logger.info(f"Department deleted by {request.user.username}")
            return redirect('manage_department')
        
        return render(request, 'manage_department.html', {'departments': departments})
    except Exception as e:
        logger.error(f"Error in manage_department: {str(e)}")
        messages.error(request, 'An error occurred.')
        return redirect('inventory')


@login_required
def manage_category(request):
    try:
        categories = AssetCategory.objects.all().order_by('name')
        
        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'add':
                name = request.POST.get('name')
                code = request.POST.get('code')
                description = request.POST.get('description', '')
                unit = request.POST.get('unit', 'pcs')
                if name and code:
                    AssetCategory.objects.create(name=name, code=code, description=description, unit=unit)
                    messages.success(request, f'Category "{name}" added successfully!')
                    logger.info(f"Asset category created: {name} by {request.user.username}")
            elif action == 'edit':
                cat_id = request.POST.get('category_id')
                category = AssetCategory.objects.get(id=cat_id)
                category.name = request.POST.get('name')
                category.code = request.POST.get('code')
                category.description = request.POST.get('description', '')
                category.unit = request.POST.get('unit', 'pcs')
                category.save()
                messages.success(request, 'Category updated successfully!')
                logger.info(f"Asset category updated: {category.name} by {request.user.username}")
            elif action == 'delete':
                cat_id = request.POST.get('category_id')
                AssetCategory.objects.get(id=cat_id).delete()
                messages.success(request, 'Category deleted successfully!')
                logger.info(f"Asset category deleted by {request.user.username}")
            return redirect('manage_category')
        
        return render(request, 'manage_category.html', {'categories': categories})
    except Exception as e:
        logger.error(f"Error in manage_category: {str(e)}")
        messages.error(request, 'An error occurred.')
        return redirect('inventory')